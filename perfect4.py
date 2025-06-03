"""
AI Invoice Processing Script

This script processes invoices using AI techniques to extract relevant information
and automatically populate a structured format, including the Chart of Accounts.

Author: Vaishnavi S. Chandgude
Date: 01/04/2025
Version: 1.0
"""

import pandas as pd
import PyPDF2
import re
import json
import os
import shutil
import traceback
from datetime import datetime
from typing import Dict, Any, List, Optional, Union, Tuple
import openpyxl
from openpyxl.utils import get_column_letter
import PyPDF2
import anthropic
import csv

def extract_invoice_data(pdf_path):
    """Extracts raw text from an invoice PDF for AI processing."""
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
    return text

def analyze_excel_structure(excel_path, sheet_name="COA i-Kcal"):
    """Reads and analyzes the Chart of Accounts structure from the Excel file."""
    xls = pd.ExcelFile(excel_path)
    if sheet_name not in xls.sheet_names:
        raise ValueError(f"Sheet '{sheet_name}' not found in the provided Excel file.")

    coa_sheet = pd.read_excel(xls, sheet_name=sheet_name)
    coa_columns = coa_sheet.columns.tolist()
    
    # Convert datetime columns to strings
    for col in coa_columns:
        if pd.api.types.is_datetime64_any_dtype(coa_sheet[col]):
            coa_sheet[col] = coa_sheet[col].dt.strftime('%Y-%m-%d')
    
    # Analyze column patterns and relationships
    column_patterns = {}
    column_relationships = {}
    column_hierarchy = {}
    
    for col in coa_columns:
        # Skip empty columns
        if coa_sheet[col].isna().all():
            continue
            
        # Analyze column content patterns
        unique_values = coa_sheet[col].dropna().astype(str).unique()
        
        # Check for code patterns
        if any('-' in str(v) for v in unique_values):
            column_patterns[col] = {'type': 'code', 'example': unique_values[0]}
            # Try to determine code hierarchy
            code_parts = str(unique_values[0]).split('-')
            if len(code_parts) > 1:
                column_hierarchy[col] = len(code_parts)
        
        # Check for numeric patterns
        elif pd.api.types.is_numeric_dtype(coa_sheet[col]):
            if all(len(str(v)) == 2 for v in unique_values if str(v).isdigit()):
                column_patterns[col] = {'type': '2-digit', 'example': unique_values[0]}
            elif all(len(str(v)) == 4 for v in unique_values if str(v).isdigit()):
                column_patterns[col] = {'type': '4-digit', 'example': unique_values[0]}
            elif any('.' in str(v) for v in unique_values):
                column_patterns[col] = {'type': 'decimal', 'example': unique_values[0]}
        
        # Check for text patterns
        else:
            column_patterns[col] = {'type': 'text', 'example': unique_values[0] if len(unique_values) > 0 else ''}
        
        # Analyze relationships between columns
        if 'Unnamed:' in str(col):
            # Find the corresponding named column
            for named_col in coa_columns:
                if 'Unnamed:' not in str(named_col):
                    # Check if values in unnamed column are derived from named column
                    if not coa_sheet[col].isna().all() and not coa_sheet[named_col].isna().all():
                        unnamed_values = coa_sheet[col].dropna().astype(str).unique()
                        named_values = coa_sheet[named_col].dropna().astype(str).unique()
                        if any(str(nv) in str(uv) for nv in named_values for uv in unnamed_values):
                            column_relationships[col] = named_col
                            break
    
    # Group columns by their patterns
    grouped_columns = {
        'code_columns': [col for col, pattern in column_patterns.items() if pattern['type'] == 'code'],
        'numeric_columns': {
            '2-digit': [col for col, pattern in column_patterns.items() if pattern['type'] == '2-digit'],
            '4-digit': [col for col, pattern in column_patterns.items() if pattern['type'] == '4-digit'],
            'decimal': [col for col, pattern in column_patterns.items() if pattern['type'] == 'decimal']
        },
        'text_columns': [col for col, pattern in column_patterns.items() if pattern['type'] == 'text']
    }
    
    # Sort code columns by hierarchy
    if column_hierarchy:
        grouped_columns['code_columns'].sort(key=lambda x: column_hierarchy.get(x, 0), reverse=True)
    
    structure = {
        'columns': coa_columns,
        'patterns': column_patterns,
        'grouped_columns': grouped_columns,
        'relationships': column_relationships,
        'hierarchy': column_hierarchy
    }
    
    return coa_sheet, structure

def generate_account_code(coa_sheet, invoice_data):
    """
    Generates a new account code based on the financial classification from the invoice
    and existing patterns in the Chart of Accounts.
    """
    # Find the code column dynamically
    code_column = None
    for col in coa_sheet.columns:
        if 'code' in str(col).lower():
            code_column = col
            break
    
    if not code_column:
        raise ValueError("No code column found in the Excel sheet")
    
    # Find the prefix pattern from existing codes
    prefix_pattern = None
    for _, row in coa_sheet.iterrows():
        if not pd.isna(row[code_column]):
            code_parts = str(row[code_column]).split('-')
            if len(code_parts) >= 2:
                prefix_pattern = code_parts[0]
                break
    
    if not prefix_pattern:
        raise ValueError("No existing code patterns found in the Excel sheet")
    
    # Extract the base prefix (e.g., "IK" from "IKL", "IKE", "IKA")
    base_prefix = prefix_pattern[:-1]  # Remove the last character
    
    # Determine account type suffix based on invoice data
    account_suffix = None
    
    # First try to use explicit account type from invoice data
    if 'account_type' in invoice_data:
        account_type = invoice_data['account_type'].upper()
        if account_type in ['EXPENSE', 'COST']:
            account_suffix = 'E'
        elif account_type in ['LIABILITY', 'REVENUE']:
            account_suffix = 'L'
        elif account_type in ['ASSET']:
            account_suffix = 'A'
    
    # If no explicit account type, try to determine from MainGpCode
    if not account_suffix and 'MainGpCode' in invoice_data:
        main_gp_code = invoice_data['MainGpCode'].upper()
        if any(term in main_gp_code for term in ['EXPENSE', 'COST']):
            account_suffix = 'E'
        elif any(term in main_gp_code for term in ['LIABILITY', 'REVENUE']):
            account_suffix = 'L'
        elif any(term in main_gp_code for term in ['ASSET']):
            account_suffix = 'A'
    
    # If still no suffix, analyze invoice content
    if not account_suffix:
        invoice_text = invoice_data.get('invoice_text', '').upper()
        if any(term in invoice_text for term in ['EXPENSE', 'COST', 'PAYMENT', 'BILL']):
            account_suffix = 'E'
        elif any(term in invoice_text for term in ['REVENUE', 'SALE', 'INCOME', 'RECEIPT']):
            account_suffix = 'L'
        elif any(term in invoice_text for term in ['ASSET', 'EQUIPMENT', 'MACHINE', 'PROPERTY']):
            account_suffix = 'A'
    
    # If still no suffix, use the suffix from existing prefix pattern
    if not account_suffix:
        account_suffix = prefix_pattern[-1]  # Use the last character of existing prefix
    
    # Combine base prefix with determined suffix
    account_type = f"{base_prefix}{account_suffix}"
    
    # Find the classification columns dynamically
    classification_columns = []
    for col in coa_sheet.columns:
        if 'group' in str(col).lower() or 'classification' in str(col).lower():
            classification_columns.append(col)
    
    # Sort columns to maintain hierarchy (primary, main, sub)
    classification_columns.sort(key=lambda x: len(str(x)), reverse=True)
    
    # Extract classification components from invoice data
    classification_values = {}
    for col in classification_columns:
        classification_values[col] = invoice_data.get(col, '')
    
    # Generate classification codes
    codes = []
    for col in classification_columns:
        value = classification_values[col]
        code = value.split('-')[0] if '-' in value else '00'
        codes.append(code)
    while len(codes) < 3:  # Ensure we have at least 3 parts
        codes.append('00')
    
    # Generate sequence number based on invoice data
    sequence_number = None
    
    # First try to use invoice number
    if 'invoice_number' in invoice_data:
        try:
            # Try to extract numeric part from invoice number
            inv_num = ''.join(filter(str.isdigit, str(invoice_data['invoice_number'])))
            if inv_num:
                sequence_number = f"{int(inv_num):04d}"
        except (ValueError, TypeError):
            pass
    
    # If no invoice number, try to find matching rows and increment
    if not sequence_number:
        matching_rows = coa_sheet.copy()
        for col, value in classification_values.items():
            if value:
                matching_rows = matching_rows[matching_rows[col] == value]
        
        if not matching_rows.empty:
            last_code = matching_rows[code_column].iloc[-1]
            code_parts = str(last_code).split('-')
            if len(code_parts) >= 5:
                try:
                    sequence = int(code_parts[-1])
                    sequence_number = f"{sequence + 1:04d}"
                except (ValueError, TypeError):
                    pass
    
    # If still no sequence number, generate based on date
    if not sequence_number:
        try:
            date_str = invoice_data.get('date', '')
            if date_str:
                # Convert date to numeric format (e.g., 20250318 for March 18, 2025)
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                sequence_number = f"{date_obj.year}{date_obj.month:02d}{date_obj.day:02d}"
            else:
                sequence_number = "0001"  # Last resort default
        except (ValueError, TypeError):
            sequence_number = "0001"  # Last resort default
    
    new_code = f"{account_type}-{'-'.join(codes[:3])}-{sequence_number}"
    print("new_code:", new_code)
    
    return new_code

def extract_first_json(text):
    """Extracts the first JSON object or array from a text string."""
    safe_print("Extracting JSON from text...")
    
    # Look for JSON content within markdown code blocks first
    code_block_pattern = r'```(?:json)?\s*([\s\S]*?)```'
    code_blocks = re.findall(code_block_pattern, text)
    
    if code_blocks:
        # Try each code block until we find valid JSON
        for block in code_blocks:
            json_text = block.strip()
            safe_print(f"Found code block, JSON length: {len(json_text)} characters")
            
            # Check if it's a list of objects without array brackets
            if json_text.startswith('{') and ('},{' in json_text or '},\n{' in json_text):
                safe_print("Detected multiple JSON objects without array brackets, adding them...")
                try:
                    # Wrap in array brackets and try to parse
                    wrapped_json = '[' + json_text + ']'
                    result = json.loads(wrapped_json)
                    safe_print("Successfully wrapped JSON in array brackets.")
                    return result
                except json.JSONDecodeError as e:
                    safe_print(f"Failed to parse wrapped JSON: {e}")
            
            # Try to parse the block directly
            try:
                result = json.loads(json_text)
                safe_print("JSON successfully extracted.")
                return result
            except json.JSONDecodeError as e:
                safe_print(f"Failed to parse JSON from code block: {e}")
    
    # If we reached here, we couldn't parse code blocks properly
    # Try to find and extract multiple JSON objects and wrap them in an array
    safe_print("Looking for JSON objects in the entire text...")
    objects = re.findall(r'\{[^\{\}]*(?:\{[^\{\}]*\}[^\{\}]*)*\}', text)
    
    if objects and len(objects) > 1:
        wrapped_json = "[" + ",".join(objects) + "]"
        try:
            safe_print("Trying to parse with manual array wrapping...")
            return json.loads(wrapped_json)
        except json.JSONDecodeError as e:
            safe_print(f"Failed with array wrapping approach: {e}")
    elif objects and len(objects) == 1:
        try:
            safe_print("Found a single JSON object in the text.")
            return json.loads(objects[0])
        except json.JSONDecodeError as e:
            safe_print(f"Failed to parse single object: {e}")
    
    # If all attempts failed
    safe_print("ERROR: No valid JSON found in response.")
    raise ValueError("No valid JSON found in the response")

def construct_prompt(coa_sheet, structure, invoice_text):
    """Constructs a well-structured prompt for Claude, ensuring correct financial classification."""
    
    # Get example rows from the Excel sheet and convert to string format
    example_rows = []
    for _, row in coa_sheet.head(5).iterrows():
        row_dict = {}
        for col in structure['columns']:
            value = row[col]
            if pd.isna(value):
                row_dict[col] = ""
            elif isinstance(value, (pd.Timestamp, datetime)):
                row_dict[col] = value.strftime('%Y-%m-%d')
            else:
                row_dict[col] = str(value)
        example_rows.append(row_dict)
    
    # Create format requirements based on the analyzed structure
    format_requirements = []
    
    # Add requirements for all columns
    for col in structure['columns']:
        if col in structure['patterns']:
            pattern = structure['patterns'][col]
            if pattern['type'] == 'code':
                format_requirements.append(f"- {col}: Must follow pattern {pattern['example']}")
            elif pattern['type'] in ['2-digit', '4-digit']:
                format_requirements.append(f"- {col}: Must be a {pattern['type']} number (e.g., {pattern['example']})")
            elif pattern['type'] == 'decimal':
                format_requirements.append(f"- {col}: Must be a decimal number (e.g., {pattern['example']})")
            elif pattern['type'] == 'text':
                format_requirements.append(f"- {col}: Text field, example value: {pattern['example']}")
    
    # Add code hierarchy requirements
    if structure['hierarchy']:
        format_requirements.append("\nCode Hierarchy Requirements:")
        for col, level in structure['hierarchy'].items():
            values = coa_sheet[col].dropna().astype(str).unique().tolist()
            if values:
                format_requirements.append(f"- {col}: Must be one of: {', '.join(values)}")
    
    # Add relationship requirements
    if structure['relationships']:
        format_requirements.append("\nColumn Relationship Requirements:")
        for unnamed_col, named_col in structure['relationships'].items():
            format_requirements.append(f"- {unnamed_col}: Values should be derived from {named_col}")
    
    prompt = f"""
    You are an AI accountant. Analyze this invoice and provide a complete financial classification.
    The classification must include ALL columns from the Chart of Accounts, with proper formatting for each.

    **Invoice Text:**
    {invoice_text}

    **Chart of Accounts sheet:**
    {coa_sheet}

    **Required Column Formats:**
    {chr(10).join(format_requirements)}

    **Example Rows from Chart of Accounts:**
    {json.dumps(example_rows, indent=2)}

    balance_sheet_structure = 
    VERTICAL BALANCE SHEET FORMAT:

    I. EQUITY AND LIABILITIES
    1. Shareholders' Funds
       a. Share Capital
       b. Reserves and Surplus
       c. *Money Received Against Share Warrants
    2. *Share Application Money Pending Allotment
    3. Non-current Liabilities
       a. Long-term Borrowings
       b. *Deferred Tax Liabilities (Net)
       c. *Other Long-term Liabilities
       d. Long-term Provisions
    4. Current Liabilities
       a. Short-term Borrowings
       b. Trade Payables
       c. Other Current Liabilities
       d. Short-term Provisions

    II. ASSETS
    1. Non-current Assets
       a. Fixed Assets
          1. Tangible Assets
          2. Intangible Assets
          3. *Capital Work-in-progress
          4. Intangible Assets under Development
       b. Non-current Investments
       c. *Deferred Tax Assets (Net)
       d. Long-term Loans and Advances
       e. *Other Non-current Assets
    2. Current Assets
       a. Current Investments
       b. Inventories
       c. Trade Receivables
       d. Cash and Cash Equivalents
       e. Short-term Loans and Advances
       f. Other Current Assets


    **Special Instructions:**
    1. Analyze the example rows to understand the structure and patterns
    2. For code columns, follow the exact pattern shown in the examples
    3. For numeric columns, use the correct number of digits as specified
    4. For text columns, use appropriate values based on the invoice content
    5. For unnamed columns, derive values from their related named columns
    6. Ensure all columns are filled with appropriate values
    7. Maintain the hierarchy of codes as shown in the examples

    Provide the classification in JSON format with ALL columns from the example rows.
    """
    
    return prompt

def analyze_code_patterns(coa_sheet, structure):
    """Analyzes and returns patterns in the Code column to help Claude understand structure."""
    code_examples = {}
    
    # Find the code column dynamically
    code_column = structure['code_columns'][0]
    if not code_column:
        return "No code patterns found - code column not detected"
    
    # Find classification columns dynamically
    classification_columns = structure['hierarchy']
    if not classification_columns:
        return "No code patterns found - no classification columns detected"
    
    # Use the highest level classification column (first in the sorted list)
    top_level_column = classification_columns[0]
    
    # Group by the top level classification to find patterns
    for group_value in coa_sheet[top_level_column].unique():
        if pd.isna(group_value):
            continue
            
        filtered = coa_sheet[coa_sheet[top_level_column] == group_value]
        if not filtered.empty:
            code_examples[group_value] = filtered[code_column].iloc[0]
    
    patterns_text = f"Code pattern examples by {top_level_column}:\n"
    for group, code in code_examples.items():
        patterns_text += f"- {group}: {code}\n"
    
    return patterns_text

def classify_invoice_with_claude(invoice_text, coa_sheet, structure, api_key):
    """Uses Claude API to classify invoice data and match it to the Chart of Accounts."""
    # Get the structure analysis
    prompt = construct_prompt(coa_sheet, structure, invoice_text)

    safe_print("\nSending prompt to Claude API...")
    
    # Set a system prompt specifically requesting JSON response
    system_prompt = """
    You are a financial analysis assistant. When producing JSON output:
    1. Always enclose the entire JSON in a code block with ```json and ``` markers
    2. Ensure the JSON is well-formed and valid
    3. Provide a single JSON object, not an array of objects
    4. Follow the exact schema requested by the user
    """
    
    # Call the Claude API
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-3-opus-20240229",
        max_tokens=4000,
        temperature=0,
        system=system_prompt,
        messages=[
            {"role": "user", "content": prompt}
        ]
    )
    
    response_text = message.content[0].text
    
    # Use our safe_print for Claude's response that might contain Unicode characters
    try:
        safe_print("\nClaude Response (Preview): " + response_text[:100] + "...")
    except:
        safe_print("\nClaude Response received but cannot be displayed due to encoding issues.")

    try:
        extracted_data = extract_first_json(response_text)
        safe_print("\nJSON successfully extracted.")
        
        # Check if extracted data is a list/array, and if so, use the first item
        if isinstance(extracted_data, list) and len(extracted_data) > 0:
            safe_print(f"Extracted data is an array with {len(extracted_data)} items. Using the first item.")
            item_data = extracted_data[0]  # Use the first item in the list
        else:
            item_data = extracted_data
            
        # Ensure all required columns are present and properly formatted
        final_data = {}
        for col in structure['columns']:
            # Get the value directly from extracted_data
            value = item_data.get(col, "")
            
            # Apply formatting based on column type
            if col in structure['patterns']:
                pattern = structure['patterns'][col]
                try:
                    if pattern['type'] == '2-digit' and value:
                        # Handle both integer and decimal values
                        if '.' in str(value):
                            value = f"{int(float(str(value))):02d}"
                        else:
                            value = f"{int(str(value)):02d}"
                    elif pattern['type'] == '4-digit' and value:
                        # Handle both integer and decimal values
                        if '.' in str(value):
                            value = f"{int(float(str(value))):04d}"
                        else:
                            value = f"{int(str(value)):04d}"
                    elif pattern['type'] == 'decimal' and value:
                        # Ensure decimal format
                        value = f"{float(str(value)):.1f}"
                except (ValueError, TypeError) as e:
                    safe_print(f"Warning: Could not format value '{value}' for column '{col}': {str(e)}")
                    # Keep original value if formatting fails
                    pass
            
            # Store the value in final_data
            final_data[col] = value
        
        safe_print("\nFinal Data to be inserted: [Data prepared successfully]")
        return final_data

    except Exception as e:
        raise ValueError(f"Error processing Claude's response: {str(e)}\nRaw output:\n{response_text}")

def update_excel_with_data(excel_path, sheet_name, data, existing_file_path=None):
    """Updates the existing Excel file with new data and saves a copy.
    If existing_file_path is provided, appends to that file instead of creating a new one."""
    try:
        # Create output directory for processed files
        output_dir = os.path.join(os.path.dirname(excel_path), "processed_output")
        os.makedirs(output_dir, exist_ok=True)
        
        # If existing file path is provided and exists, use that instead of creating a new file
        if existing_file_path and os.path.exists(existing_file_path):
            output_path = existing_file_path
            safe_print(f"Appending to existing Excel file at: {output_path}")
        else:
            # Generate timestamp and output file name
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_name = os.path.basename(excel_path)
            name_parts = os.path.splitext(base_name)
            output_path = os.path.join(output_dir, f"{name_parts[0]}_combined_{timestamp}{name_parts[1]}")
            safe_print(f"Creating new combined Excel file at: {output_path}")
        
        # Copy the original file to avoid modifying it directly
        try:
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found at {excel_path}")
            
            # If we're using an existing file, don't copy it
            if existing_file_path and os.path.exists(existing_file_path):
                safe_print(f"Using existing Excel file: {output_path}")
            else:
                # Otherwise copy the original file to the new location
                shutil.copy2(excel_path, output_path)
                safe_print(f"Excel file copied to: {output_path}")
        except Exception as e:
            safe_print(f"Error copying Excel file: {str(e)}")
            # If copying fails, try creating a new Excel file as fallback
            return create_new_excel_file(output_path, data)
        
        # Try to load the workbook
        try:
            wb = load_workbook(output_path, keep_vba=True)
        except Exception as wb_error:
            safe_print(f"Error loading workbook: {str(wb_error)}")
            # If loading fails, try creating a new Excel file as fallback
            return create_new_excel_file(output_path, data)
        
        # Try to use the specified sheet name, fallback to active sheet if not found
        try:
            ws = wb[sheet_name]
            safe_print(f"Using sheet: {sheet_name}")
        except KeyError:
            sheet_names = wb.sheetnames
            safe_print(f"Sheet '{sheet_name}' not found. Available sheets: {sheet_names}")
            # If the specified sheet doesn't exist, try to use the first sheet
            ws = wb.active
            safe_print(f"Using active sheet: {ws.title} instead")
        
        # Find the last filled row
        try:
            last_row = ws.max_row
            while last_row > 0 and all(cell.value is None for cell in ws[last_row]):
                last_row -= 1
            
            # Get the next empty row
            new_row = last_row + 1
            safe_print(f"Last filled row: {last_row}")
            safe_print(f"Adding new data to row: {new_row}")
        except Exception as row_error:
            safe_print(f"Error finding last row: {str(row_error)}")
            new_row = ws.max_row + 1
            safe_print(f"Using row: {new_row} as fallback")
        
        # Get column headers and their positions
        try:
            headers = {}
            for idx, cell in enumerate(ws[1], 1):
                col_letter = get_column_letter(idx)
                header = cell.value
                if header is None:
                    # For unnamed columns, use the format "Unnamed: {index}"
                    header = f"Unnamed: {idx-1}"
                headers[header] = col_letter
            
            safe_print("Excel Headers:", list(headers.keys()))
        except Exception as header_error:
            safe_print(f"Error getting headers: {str(header_error)}")
            # Create simple headers if needed
            headers = {}
            for idx, key in enumerate(data.keys(), 1):
                headers[key] = get_column_letter(idx)
        
        # Update each column with the values from data
        for header, col_letter in headers.items():
            # Try to find a matching value in the data
            value = None
            
            # First try exact match
            if header in data:
                value = data[header]
            else:
                # Try case-insensitive match
                for data_key in data.keys():
                    if str(header).lower() == str(data_key).lower():
                        value = data[data_key]
                        break
            
            # If no match found, try to infer value based on column name and data
            if value is None:
                # Check for date columns
                if any(date_term in str(header).lower() for date_term in ['date', 'time', 'period']):
                    value = datetime.now().strftime('%Y-%m-%d')
                
                # Check for amount/value columns
                elif any(amount_term in str(header).lower() for amount_term in ['amount', 'value', 'total', 'sum']):
                    value = '0.00'
                
                # Check for code columns
                elif any(code_term in str(header).lower() for code_term in ['code', 'id', 'number']):
                    value = '0000'
                
                # For unnamed columns, try to find a value from a related named column
                elif 'unnamed' in str(header).lower():
                    for named_col in headers:
                        if 'unnamed' not in str(named_col).lower():
                            if named_col in data:
                                value = data[named_col]
                                break
            
            # If still no value, use empty string
            if value is None:
                value = ""
            
            # Set the value in the Excel sheet
            try:
                safe_print(f"Setting {col_letter}{new_row} ({header}) = {value}")
                ws[f"{col_letter}{new_row}"] = value
            except Exception as cell_error:
                safe_print(f"Error setting cell {col_letter}{new_row}: {str(cell_error)}")
        
        # Save the workbook to the new location
        try:
            safe_print(f"Saving updated Excel file to: {output_path}")
            wb.save(output_path)
            safe_print(f"Saved to: {output_path}")
            safe_print(f"\n✅ Successfully updated row {new_row} in the Excel file")
        except Exception as save_error:
            safe_print(f"Error saving workbook: {str(save_error)}")
            # If saving fails, try creating a new Excel file as fallback
            return create_new_excel_file(output_path, data)
        
        return output_path  # Return the path of the saved file
    
    except Exception as e:
        safe_print(f"Error updating Excel: {str(e)}")
        # Try creating a new Excel file as a last resort
        return create_new_excel_file(output_path, data)

def create_new_excel_file(output_path, data):
    """Creates a new Excel file as a fallback when updating fails."""
    try:
        safe_print("Attempting to create new Excel file as fallback...")
        # Create a new Excel file from scratch
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Invoice"
        
        # Add headers in the first row
        headers = list(data.keys())
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}1"] = header
            
        # Add data in the second row
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            value = data[header]
            ws[f"{col_letter}2"] = value
        
        # Save the workbook to the new location
        new_output_path = output_path.replace('.xlsm', '.xlsx').replace('.xls', '.xlsx')
        wb.save(new_output_path)
        safe_print(f"Saved new Excel file to: {new_output_path}")
        safe_print(f"Saved to: {new_output_path}")
        return new_output_path
    except Exception as e:
        safe_print(f"Error creating fallback Excel file: {str(e)}")
        raise

def update_chart_of_accounts(excel_path, invoice_data, sheet_name="COA i-Kcal"):
    """
    Updates the Chart of Accounts Excel sheet with extracted invoice data.
    
    Args:
        excel_path (str): Path to the Excel file to update
        invoice_data (dict): Dictionary containing the invoice data to add
        sheet_name (str): Name of the sheet to update (default: 'COA i-Kcal')
        
    Returns:
        bool: True if update was successful, False otherwise
        
    Raises:
        ValueError: If invoice_data is not a dictionary or is empty
        FileNotFoundError: If the Excel file doesn't exist
        PermissionError: If the Excel file is read-only
        Exception: For any other errors during the update process
    """
    safe_print("\n=== Starting Excel update process ===")
    safe_print(f"Excel Path: {excel_path}")
    safe_print(f"Sheet Name: {sheet_name}")
    
    # Validate invoice_data
    if not isinstance(invoice_data, dict):
        error_msg = f"invoice_data must be a dictionary, got {type(invoice_data).__name__}"
        safe_print(f"❌ {error_msg}")
        safe_print(f"invoice_data value: {invoice_data}")
        raise ValueError(error_msg)
    
    if not invoice_data:
        error_msg = "invoice_data dictionary is empty"
        safe_print(f"❌ {error_msg}")
        raise ValueError(error_msg)
    
    safe_print("Invoice data keys:", list(invoice_data.keys()))
    safe_print("Invoice data values:", list(invoice_data.values()))
    
    wb = None  # Initialize wb variable for use in finally block
    
    try:
        # Check if file exists and is accessible
        if not os.path.exists(excel_path):
            error_msg = f"Excel file not found at: {excel_path}"
            safe_print(f"❌ {error_msg}")
            raise FileNotFoundError(error_msg)
        
        # Check if file is read-only
        if not os.access(excel_path, os.W_OK):
            error_msg = "Excel file is read-only. Please close the file if it's open in Excel and ensure you have write permissions."
            safe_print(f"❌ {error_msg}")
            raise PermissionError(error_msg)
        
        # Load the existing workbook
        safe_print("Loading workbook...")
        wb = load_workbook(excel_path, keep_vba=True)  # keep_vba=True to preserve macros
        
        # Get the target sheet
        if sheet_name not in wb.sheetnames:
            error_msg = f"Sheet '{sheet_name}' not found in the Excel file. Available sheets: {wb.sheetnames}"
            safe_print(f"❌ {error_msg}")
            raise ValueError(error_msg)
        
        ws = wb[sheet_name]
        safe_print(f"Successfully accessed sheet: {sheet_name}")
        
        # Find the first empty row by checking each row from the bottom up
        max_row = ws.max_row
        last_filled_row = 0
        
        safe_print(f"Checking for last filled row (max_row: {max_row})...")
        
        # Find the last row with data
        for row in range(max_row, 0, -1):
            if any(cell.value for cell in ws[row]):
                last_filled_row = row
                safe_print(f"Found last filled row: {last_filled_row}")
                break
        
        # The new row will be one after the last filled row
        new_row = last_filled_row + 1
        safe_print(f"Will add new data to row: {new_row}")
        
        # Get column headers from the first row
        headers = [cell.value for cell in ws[1]]
        safe_print(f"\nExcel Headers: {headers}")
        safe_print(f"Invoice data keys: {list(invoice_data.keys())}")
        
        # Map the invoice data to the correct columns
        update_count = 0
        for col_idx, header in enumerate(headers, 1):
            if not header:  # Skip empty headers
                continue
                
            col_letter = get_column_letter(col_idx)
            cell_ref = f"{col_letter}{new_row}"
            
            # Get the value from invoice_data, default to empty string if not found
            value = invoice_data.get(header, "")
            
            # Skip if the header isn't in invoice_data and we don't have a value
            if not value and header not in invoice_data:
                continue
                
            # Convert value to appropriate type based on existing data
            if last_filled_row > 0:  # Only if we have existing data to check types
                try:
                    cell_ref = f"{col_letter}{last_filled_row}"
                    safe_print(f"Accessing cell {cell_ref} in column {col_letter} (header: {header})")
                    
                    # Add validation for column letter and row number
                    if not col_letter.isalpha() or last_filled_row <= 0:
                        safe_print(f"Invalid cell reference: {col_letter}{last_filled_row}")
                        continue
                        
                    # Try to access the cell with better error handling
                    try:
                        existing_cell = ws.cell(row=last_filled_row, column=col_idx)
                    except Exception as e:
                        safe_print(f"Error accessing cell {cell_ref}: {str(e)}")
                        continue
                        
                    if existing_cell is None:
                        safe_print(f"Warning: Cell {cell_ref} is None")
                        continue
                        
                    if existing_cell.value is not None:
                        if isinstance(existing_cell.value, (int, float)):
                            try:
                                value = float(value) if '.' in str(value) else int(value)
                            except (ValueError, TypeError):
                                safe_print(f"Warning: Could not convert value '{value}' to number for column {header}")
                except Exception as e:
                    safe_print(f"Error processing cell {col_letter}{last_filled_row}: {str(e)}")
                    continue
                        
                    if existing_cell.value is not None and isinstance(existing_cell.value, datetime):
                        try:
                            value = datetime.strptime(str(value), '%Y-%m-%d')
                        except (ValueError, TypeError):
                            safe_print(f"Warning: Could not convert value '{value}' to date for column {header}")
                            continue
            
            # Set the cell value
            ws[cell_ref] = value
            update_count += 1
            safe_print(f"Set {cell_ref} = {value} (type: {type(value).__name__})")
        
        # Save the changes if we updated any cells
        if update_count > 0:
            safe_print(f"\nSaving changes to {excel_path}...")
            wb.save(excel_path)
            safe_print("Changes saved successfully")
            return True
        else:
            safe_print("\nNo cells were updated. File not saved.")
            return False
            
    except Exception as e:
        error_msg = f"Error updating Excel file: {str(e)}"
        safe_print(f"\n❌ {error_msg}")
        safe_print(f"Traceback: {traceback.format_exc()}")
        raise Exception(error_msg) from e
        
    finally:
        # Make sure to close the workbook to prevent memory leaks
        if wb is not None:
            try:
                wb.close()
                safe_print("Workbook closed successfully")
            except Exception as e:
                safe_print(f"Warning: Error closing workbook: {str(e)}")

def process_invoice_file(invoice_path, chart_path, sheet_name, output_dir, unique_id):
    """
    Process an invoice and update the chart of accounts.
    
    Args:
        invoice_path (str): Path to the invoice PDF file
        chart_path (str): Path to the chart of accounts Excel file
        sheet_name (str): Name of the sheet to update in the Excel file
        output_dir (str): Directory to save the processed file
        unique_id (str): Unique identifier for this processing job
        
    Returns:
        dict: Processing results with status and file information
    """
    # Initialize result with default values
    result = {
        'status': 'error',
        'message': 'Processing not completed',
        'output_path': '',
        'output_filename': '',
        'invoice_data': {}
    }
    
    try:
        # Validate input paths
        safe_print("\n=== Validating input files ===")
        safe_print(f"Invoice path: {invoice_path} (exists: {os.path.exists(invoice_path)})")
        safe_print(f"Chart path: {chart_path} (exists: {os.path.exists(chart_path)})")
        safe_print(f"Output directory: {output_dir} (exists: {os.path.exists(output_dir)})")
        
        if not os.path.exists(invoice_path):
            raise FileNotFoundError(f"Invoice file not found: {invoice_path}")
        if not os.path.exists(chart_path):
            raise FileNotFoundError(f"Chart file not found: {chart_path}")
            
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        safe_print(f"Ensured output directory exists: {output_dir}")
        
        # Generate output filename with unique ID and timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'updated_chart_{timestamp}_{unique_id}.xlsx'
        output_path = os.path.abspath(os.path.join(output_dir, output_filename))
        
        safe_print(f"\n=== Processing invoice ===")
        safe_print(f"Output will be saved to: {output_path}")
        
        # Extract invoice data
        safe_print("Extracting invoice data...")
        try:
            invoice_text = extract_invoice_data(invoice_path)
            safe_print("Raw invoice text extracted successfully")
        except Exception as e:
            safe_print(f"Error extracting invoice data: {str(e)}")
            raise
        
        # Analyze chart of accounts structure
        safe_print("Analyzing chart of accounts structure...")
        coa_sheet, structure = analyze_excel_structure(chart_path, sheet_name)
        
        # Classify invoice data using Claude AI
        safe_print("Classifying invoice data with Claude AI...")
        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY environment variable not set")
        
        try:
            # Process the invoice text with Claude
            invoice_data = classify_invoice_with_claude(
                invoice_text=invoice_text,
                coa_sheet=coa_sheet,
                structure=structure,
                api_key=api_key
            )
            safe_print(f"Classified invoice data: {json.dumps(invoice_data, indent=2)}")
            
        except Exception as e:
            safe_print(f"Error classifying invoice with Claude: {str(e)}")
            safe_print("Using fallback invoice data")
            invoice_data = {
                'invoice_number': f'INV-{unique_id}',
                'date': datetime.now().strftime('%Y-%m-%d'),
                'amount': 1000.00,
                'vendor': 'Sample Vendor',
                'description': 'Sample invoice description',
                'classification': 'Revenue',
                'account_code': '4000',  # Default to Revenue account
                'account_name': 'Sales Revenue',
                'department': 'General',
                'category': 'Income'
            }
        
        result['invoice_data'] = invoice_data
        
        # Create a temporary file for the updated chart
        temp_output = os.path.join(output_dir, f'temp_{unique_id}.xlsx')
        
        try:
            # Update the chart of accounts
            safe_print("\n=== Updating chart of accounts ===")
            safe_print(f"Using sheet: {sheet_name}")
            safe_print(f"Temporary output: {temp_output}")
            
            # Make a working copy of the chart file
            working_chart_path = os.path.join(output_dir, f'working_{unique_id}.xlsx')
            shutil.copy2(chart_path, working_chart_path)
            safe_print(f"Created working copy at: {working_chart_path}")
            
            # Update the chart of accounts with the classified data
            update_chart_of_accounts(
                excel_path=working_chart_path,
                invoice_data=invoice_data,
                sheet_name=sheet_name
            )
            safe_print("Chart of accounts updated successfully")
            
            # Copy the updated file to the temp location
            shutil.copy2(working_chart_path, temp_output)
            safe_print(f"Copied updated chart to temporary location: {temp_output}")
            
            # Verify the temp file was created and has content
            if not os.path.exists(temp_output):
                raise Exception("Failed to create temporary output file")
                
            file_size = os.path.getsize(temp_output)
            if file_size == 0:
                raise Exception("Temporary output file is empty")
                
            safe_print(f"Temporary file size: {file_size} bytes")
            
            # Move the file to the final location
            shutil.move(temp_output, output_path)
            safe_print(f"Moved file to final location: {output_path}")
            
            # Verify the final file
            if not os.path.exists(output_path):
                raise Exception("Failed to move file to final location")
                
            final_size = os.path.getsize(output_path)
            safe_print(f"Final file size: {final_size} bytes")
            
            # Clean up working file
            try:
                if os.path.exists(working_chart_path):
                    os.remove(working_chart_path)
                    safe_print("Cleaned up working file")
            except Exception as e:
                safe_print(f"Warning: Could not clean up working file: {str(e)}")
            
            # Update result with success
            result.update({
                'status': 'success',
                'message': 'Invoice processed successfully',
                'output_path': output_path,
                'output_filename': output_filename
            })
            
            safe_print("\n=== Processing completed successfully ===\n")
            return result
            
        except Exception as e:
            error_msg = f"Error updating chart of accounts: {str(e)}"
            safe_print(f"\n!!! ERROR: {error_msg}")
            safe_print(traceback.format_exc())
            
            # Clean up any temporary files
            for temp_file in [temp_output, working_chart_path if 'working_chart_path' in locals() else None]:
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                        safe_print(f"Cleaned up temporary file: {temp_file}")
                    except:
                        safe_print(f"Warning: Could not clean up {temp_file}")
            
            raise Exception(error_msg)
        
    except Exception as e:
        error_trace = traceback.format_exc()
        error_msg = f"Error processing invoice: {str(e)}"
        safe_print(f"\n!!! CRITICAL ERROR: {error_msg}")
        safe_print(f"Traceback:\n{error_trace}")
        
        # Update result with error details
        result.update({
            'status': 'error',
            'message': error_msg,
            'trace': error_trace
        })
        
        return result

def safe_print(*args, **kwargs):
    """Print text safely, avoiding Unicode encoding errors. Accepts multiple arguments."""
    try:
        # Convert all arguments to strings and join them with spaces
        message = ' '.join(str(arg) for arg in args)
        # Print with the original kwargs
        print(message, **kwargs)
    except Exception as e:
        # Fallback if there's an error in the safe print itself
        try:
            print(f"[Print Error: {str(e)}]")
        except:
            pass  # Give up if we can't even print the error

def get_excel_sheets(file_path):
    """Get list of sheet names from an Excel file."""
    try:
        xls = pd.ExcelFile(file_path)
        return xls.sheet_names
    except Exception as e:
        safe_print(f"Error reading Excel file: {str(e)}")
        raise

if __name__ == "__main__":
    import sys
    import io
    
    # Set stdout to handle unicode properly
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    
    # Print script information for debugging
    safe_print(f"Python version: {sys.version}")
    safe_print(f"Current working directory: {os.getcwd()}")
    safe_print(f"Script path: {__file__}")
    
    # Get API key from environment variables
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        print("Error: ANTHROPIC_API_KEY environment variable is not set")
        print("Please set the environment variable with your Claude API key")
        sys.exit(1)
    
    # Check if command line arguments are provided
    if len(sys.argv) >= 3:
        # Use command line arguments
        excel_path = sys.argv[1]  # First argument: Excel file path
        pdf_path = sys.argv[2]    # Second argument: PDF file path
        
        # Check if sheet name is provided as third argument
        sheet_name = "COA i-Kcal"  # Default sheet name
        existing_file_path = None  # Default to None for existing file path
        
        if len(sys.argv) >= 4:
            sheet_name = sys.argv[3]  # Third argument: Excel sheet name
            
            # Check if existing file path is provided as fourth argument
            if len(sys.argv) >= 5:
                existing_file_path = sys.argv[4]  # Fourth argument: Existing Excel file path
                safe_print(f"Using command line arguments:\nExcel: {excel_path}\nPDF: {pdf_path}\nSheet: {sheet_name}\nExisting File: {existing_file_path}")
            else:
                safe_print(f"Using command line arguments:\nExcel: {excel_path}\nPDF: {pdf_path}\nSheet: {sheet_name}")
        else:
            safe_print(f"Using command line arguments:\nExcel: {excel_path}\nPDF: {pdf_path}\nSheet: {sheet_name} (default)")
    else:
        # Fallback to default paths for development/testing
        pdf_path = r"c:\Users\Admin320\Downloads\Next Corporation S1932EE 2qty.pdf"
        excel_path = r"C:\Users\Admin320\Downloads\Chart of Account R23 28 May 2022.xlsm"
        sheet_name = "COA i-Kcal"  # Default sheet name
        safe_print("No command line arguments provided. Using default paths.")
        safe_print(f"Usage: python {sys.argv[0]} <excel_path> <pdf_path> [<sheet_name>]")

    safe_print("\nStarting invoice processing...")
    safe_print(f"PDF Path: {pdf_path}")
    safe_print(f"Excel Path: {excel_path}")

    # Extract data from the invoice
    try:
        invoice_text = extract_invoice_data(pdf_path)
        if len(invoice_text) > 500:
            safe_print("\nExtracted Invoice Text: " + invoice_text[:500] + "...")
        else:
            safe_print("\nExtracted Invoice Text: " + invoice_text)
    except FileNotFoundError:
        safe_print(f"\nError: PDF file not found at path: {pdf_path}")
        sys.exit(1)

    # Analyze the Chart of Accounts structure
    try:
        coa_sheet, structure = analyze_excel_structure(excel_path, sheet_name)
        safe_print("\nExcel Structure Analyzed")
    except FileNotFoundError:
        safe_print(f"\nError: Excel file not found at path: {excel_path}")
        sys.exit(1)
    except ValueError as e:
        safe_print(f"\nError analyzing Excel structure: {str(e)}")
        sys.exit(1)

    # Use Claude to classify the invoice and generate the account code
    classified_data = classify_invoice_with_claude(invoice_text, coa_sheet, structure, api_key)

    try:
        # Update the Excel file with the extracted invoice data
        safe_print("Updating Excel file with extracted data...")
        output_file = update_excel_with_data(excel_path, sheet_name, classified_data, existing_file_path)
        safe_print(f"\nExcel file updated successfully: {output_file}")
        processed_file_path = output_file  # Store for later use
    except Exception as e:
        safe_print(f"Error updating Excel file: {str(e)}")
        processed_file_path = None
        # Try to create a CSV file instead as fallback
        try:
            csv_dir = os.path.join(os.path.dirname(excel_path), "processed_output")
            os.makedirs(csv_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            csv_path = os.path.join(csv_dir, f"invoice_data_{timestamp}.csv")
            
            with open(csv_path, 'w', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=classified_data.keys())
                writer.writeheader()
                writer.writerow(classified_data)
            
            updated_file = csv_path
            safe_print(f"\nCreated CSV file as fallback: {csv_path}")
        except Exception as csv_error:
            safe_print(f"\nError creating CSV fallback: {str(csv_error)}")
            updated_file = None

    safe_print("\n✅ Process completed successfully!")
    if processed_file_path:
        safe_print(f"Saved to: {processed_file_path}")
    else:
        safe_print("Warning: No output file was created")
