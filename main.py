from flask import Flask, jsonify, request, send_from_directory, current_app
import os
import sys
import traceback
from datetime import datetime
import uuid
from flask_cors import CORS
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import PyPDF2
from anthropic import Anthropic
import json
import re
from perfect4 import (
    extract_invoice_data,
    analyze_excel_structure,
    classify_invoice_with_claude,
    update_chart_of_accounts,
    safe_print
)

# Initialize Anthropic client
anthropic_client = Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))

# Helper function to analyze chart of accounts with Claude
def analyze_coa_with_claude(coa_data):
    try:
        # Get example rows from the Excel sheet and convert to string format
        example_rows = []
        for _, row in coa_data.head(5).iterrows():
            row_dict = {}
            for col in coa_data.columns:
                value = row[col]
                if pd.isna(value):
                    row_dict[col] = ""
                elif isinstance(value, (pd.Timestamp, datetime)):
                    row_dict[col] = value.strftime('%Y-%m-%d')
                else:
                    row_dict[col] = str(value)
            example_rows.append(row_dict)

        # Analyze patterns in code columns
        code_patterns = {}
        for col in coa_data.columns:
            if 'code' in col.lower() or 'account' in col.lower():
                patterns = coa_data[col].dropna().astype(str).unique().tolist()[:5]
                if patterns:
                    code_patterns[col] = patterns

        # Analyze hierarchy and relationships
        hierarchy = {}
        relationships = {}
        for col in coa_data.columns:
            if 'group' in col.lower() or 'category' in col.lower():
                values = coa_data[col].dropna().astype(str).unique().tolist()
                if values:
                    hierarchy[col] = values
            if 'code' in col.lower() and any(ref in col.lower() for ref in ['ref', 'related', 'parent']):
                related_cols = [c for c in coa_data.columns if c != col and 'name' in c.lower()]
                if related_cols:
                    relationships[col] = related_cols[0]

        # Prepare the prompt for Claude
        prompt = f"""You are an AI accountant. Analyze this chart of accounts and provide a complete financial classification structure.

**Chart of Accounts Data:**
{coa_data.to_string()}

**Example Rows:**
{json.dumps(example_rows, indent=2)}

**Code Patterns Found:**
{json.dumps(code_patterns, indent=2)}

**Hierarchy Information:**
{json.dumps(hierarchy, indent=2)}

**Column Relationships:**
{json.dumps(relationships, indent=2)}

Please analyze and provide:
1. The complete account hierarchy (main groups, sub groups, detail accounts)
2. The account types and their purposes
3. Specific patterns in account codes and their meanings
4. Rules for invoice categorization based on:
   - Account code structure
   - Group hierarchies
   - Naming conventions
   - Common transaction types

Provide your analysis in this JSON format:
{{
    "hierarchy": {{
        "main_groups": [],
        "sub_groups": {{}},
        "detail_accounts": {{}}
    }},
    "account_types": {{
        "asset": [],
        "liability": [],
        "equity": [],
        "revenue": [],
        "expense": []
    }},
    "code_patterns": {{
        "structure": "",
        "examples": {{}}
    }},
    "categorization_rules": [
        {{
            "pattern": "",
            "account_type": "",
            "description": ""
        }}
    ]
}}"""

        # Get Claude's analysis
        message = anthropic.messages.create(
            model="claude-3-opus-20240229",
            max_tokens=4000,
            temperature=0,
            system="You are an expert accountant specializing in financial analysis and chart of accounts structure. Always provide detailed, structured analysis in the exact JSON format requested.",
            messages=[
                {"role": "user", "content": prompt}
            ]
        )

        # Extract JSON from response
        match = re.search(r'```json\s*({[\s\S]*?})\s*```', message.content[0].text)
        if not match:
            raise ValueError("No JSON found in Claude's response")

        analysis = json.loads(match.group(1))
        return analysis

    except Exception as e:
        print(f"Error analyzing chart of accounts with Claude: {str(e)}")
        return {"error": str(e)}

# Load environment variables from .env file if it exists
load_dotenv()

# Get base directory for file storage
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Define folder paths
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', os.path.join(BASE_DIR, 'uploads'))
PROCESSED_FOLDER = os.environ.get('PROCESSED_FOLDER', os.path.join(BASE_DIR, 'processed'))
TEMP_FOLDER = os.environ.get('TEMP_FOLDER', os.path.join(BASE_DIR, 'temp'))

# Create required folders immediately
for folder in [UPLOAD_FOLDER, TEMP_FOLDER, PROCESSED_FOLDER]:
    os.makedirs(folder, exist_ok=True)
    print(f"Created directory: {folder}")

# Create the Flask application
app = Flask(__name__)

# Enable CORS for all routes and origins
CORS(app)

# Configure file upload settings
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER
app.config['TEMP_FOLDER'] = TEMP_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# In Flask 2.3+, before_first_request is removed
# Instead, we'll create a function that runs with the first request
@app.before_request
def initialize_app():
    # Only run once using an app variable to track initialization
    if not getattr(app, 'initialized', False):
        app.initialized = True
        # Any additional initialization can go here

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "message": "Flask API is running"
    })

@app.route('/', methods=['GET'])
def index():
    return """
    <html>
        <head>
            <title>Invoice Processor API</title>
            <style>
                body { font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }
                h1 { color: #333; }
                .endpoint { background: #f4f4f4; padding: 10px; margin: 10px 0; border-radius: 5px; }
                code { background: #e4e4e4; padding: 2px 5px; border-radius: 3px; }
            </style>
        </head>
        <body>
            <h1>Invoice Processor API Documentation</h1>
            <p>This API allows you to process invoices against a chart of accounts using Claude AI.</p>
            
            <h2>API Endpoints</h2>
            
            <div class="endpoint">
                <h3>1. Health Check</h3>
                <code>GET /api/health</code>
                <p>Returns the API status and timestamp.</p>
            </div>
            
            <div class="endpoint">
                <h3>2. Process Invoice</h3>
                <code>POST /api/process-invoice</code>
                <p>Processes an invoice against a chart of accounts.</p>
            </div>
            
            <div class="endpoint">
                <h3>3. Get Excel Sheets</h3>
                <code>POST /api/get-sheets</code>
                <p>Returns the sheet names from an Excel file.</p>
            </div>
            
            <div class="endpoint">
                <h3>4. Download File</h3>
                <code>GET /api/download-file/{filename}</code>
                <p>Downloads a processed file.</p>
            </div>
        </body>
    </html>
    """

# This is what Render.com needs - app must be importable from this file
application = app

# Route to handle file uploads and process invoices
@app.route('/api/process-invoice', methods=['POST'])
def process_invoice():
    try:
        # Log detailed information about the request for debugging
        print("Received request files:", list(request.files.keys()))
        print("Received form data:", list(request.form.keys()))
        print("Request content type:", request.content_type)
        print("Request data size:", request.content_length)
        
        # Get sheet name from form data
        sheet_name = request.form.get('sheetName', '')
        
        # Check if we have a chart of accounts file
        coa_file_content = request.form.get('coaFile', '')
        invoice_file_content = request.form.get('invoiceFile', '')
        
        # Generate a unique filename for this request
        unique_id = str(uuid.uuid4())[:8]
        result_filename = f"processed_invoice_{unique_id}.xlsx"
        result_path = os.path.join(app.config['PROCESSED_FOLDER'], result_filename)
        
        # Try to load the chart of accounts data if provided as base64 content
        coa_data = None
        if request.files and 'coaFile' in request.files:
            print("Processing uploaded chart of accounts file")
            coa_file = request.files['coaFile']
            if coa_file and coa_file.filename.endswith(('.xlsx', '.xls')):
                # Save the file temporarily
                temp_coa_path = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_coa_{unique_id}.xlsx")
                coa_file.save(temp_coa_path)
                
                # Try to read the Excel file
                try:
                    # If sheet name is provided, use it, otherwise read the first sheet
                    if sheet_name:
                        coa_data = pd.read_excel(temp_coa_path, sheet_name=sheet_name)
                    else:
                        coa_data = pd.read_excel(temp_coa_path)
                    print(f"Successfully read chart of accounts with {len(coa_data)} rows")
                except Exception as e:
                    print(f"Error reading chart of accounts Excel: {str(e)}")
                    # Still create output but note the error
                    coa_data = pd.DataFrame({"Error": [f"Could not read chart of accounts: {str(e)}"]})  
        
        # Create the output Excel with both request metadata and chart of accounts data if available
        # Create metadata sheet
        metadata = {
            'Timestamp': [datetime.now().isoformat()],
            'Request Content Type': [request.content_type],
            'Received Form Data': [', '.join(list(request.form.keys()))],
            'Invoice File': [invoice_file_content or 'No invoice file provided'],
            'Chart of Accounts File': [coa_file_content or 'No chart of accounts file provided'],
            'Sheet Name': [sheet_name or 'No sheet name provided']
        }
        
        # Create an Excel writer to save multiple sheets
        with pd.ExcelWriter(result_path, engine='openpyxl') as writer:
            # Write the metadata sheet
            metadata_df = pd.DataFrame(metadata)
            metadata_df.to_excel(writer, sheet_name='Request Info', index=False)
        # Check if files are present in the request
        if 'invoiceFile' not in request.files or 'coaFile' not in request.files:
            return jsonify({
                'error': 'Both invoice and chart of accounts files are required'
            }), 400
            
        # Get files from the request
        invoice_file = request.files['invoiceFile']
        chart_file = request.files['coaFile']
        sheet_name = request.form.get('sheetName', '')  # Optional sheet name from frontend
        
        # Check if files are selected
        if invoice_file.filename == '' or chart_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        try:
            # Generate unique filenames to prevent overwriting
            unique_id = str(uuid.uuid4())
            
            # Secure the filenames
            invoice_filename = secure_filename(invoice_file.filename)
            chart_filename = secure_filename(chart_file.filename)
            
            # Create file paths
            invoice_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{unique_id}_{invoice_filename}")
            chart_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{unique_id}_{chart_filename}")
            
            # Save the uploaded files
            invoice_file.save(invoice_path)
            chart_file.save(chart_path)
            
            # Process the files using perfect4.py logic
            result = process_files(invoice_path, chart_path, sheet_name, unique_id)
            
            if result.get('status') == 'success':
                return jsonify({
                    'status': 'success',
                    'message': 'Invoice and chart of accounts processed successfully',
                    'file_info': {
                        'path': result['output_path'],
                        'filename': result['filename'],
                        'download_url': f'/api/download-file?filename={result["filename"]}',
                        'file_type': 'excel'
                    },
                    'received_form_data': list(request.form.keys())
                })
            else:
                # Clean up uploaded files if processing failed
                if os.path.exists(invoice_path):
                    os.remove(invoice_path)
                if os.path.exists(chart_path):
                    os.remove(chart_path)
                    
                return jsonify({
                    'error': result.get('message', 'Failed to process invoice')
                }), 500
                
        except Exception as e:
            # Clean up any partially uploaded files if an error occurs
            if 'invoice_path' in locals() and os.path.exists(invoice_path):
                os.remove(invoice_path)
            if 'chart_path' in locals() and os.path.exists(chart_path):
                os.remove(chart_path)
                
            error_message = f"Error processing request: {str(e)}"
            print(error_message)
            traceback.print_exc()
            return jsonify({'error': error_message}), 500
            
    except Exception as e:
        print(f"Error processing invoice: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'error': f"An error occurred: {str(e)}"
        }), 500

# Function to process the files using the perfect4.py logic
def process_files(invoice_path, chart_path, sheet_name, unique_id):
    try:
        # 1. Extract text from the invoice PDF
        invoice_text = extract_invoice_data(invoice_path)
        if not invoice_text:
            raise ValueError("Could not extract text from invoice PDF")

        # 2. Analyze the Chart of Accounts Excel structure
        excel_structure = analyze_excel_structure(chart_path, sheet_name)
        if not excel_structure:
            raise ValueError("Could not analyze Chart of Accounts structure")

        # 3. Read the Chart of Accounts data
        coa_sheet = pd.read_excel(chart_path, sheet_name=sheet_name)

        # 4. Use Claude to classify invoice and match to Chart of Accounts
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY not found in environment variables")

        invoice_data = classify_invoice_with_claude(invoice_text, coa_sheet, excel_structure, api_key)
        if not invoice_data or 'error' in invoice_data:
            raise ValueError(f"Claude classification failed: {invoice_data.get('error', 'Unknown error')}")

        # 5. Create the output Excel file with multiple sheets
        output_filename = f"{unique_id}_processed_results.xlsx"
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)

        # Create a new workbook
        wb = Workbook()

        # Sheet 1: Request Info
        ws1 = wb.active
        ws1.title = "Request Info"
        request_info = [
            ["Timestamp", datetime.now().isoformat()],
            ["Invoice File", os.path.basename(invoice_path)],
            ["Chart of Accounts File", os.path.basename(chart_path)],
            ["Sheet Name", sheet_name],
            ["Processing ID", unique_id]
        ]
        for row in request_info:
            ws1.append(row)

        # Sheet 2: Chart of Accounts Data
        ws2 = wb.create_sheet("Chart of Accounts")
        for r_idx, row in enumerate(dataframe_to_rows(coa_sheet, index=False, header=True)):
            ws2.append(row)

        # Sheet 3: Processed Invoice Data
        ws3 = wb.create_sheet("Processed Invoice")
        headers = ["Date", "Description", "Amount", "Account Code", "Account Name", "Classification"]
        ws3.append(headers)

        # Add the processed invoice data
        for entry in invoice_data.get('entries', []):
            row = [
                entry.get('date', ''),
                entry.get('description', ''),
                entry.get('amount', ''),
                entry.get('account_code', ''),
                entry.get('account_name', ''),
                entry.get('classification', '')
            ]
            ws3.append(row)

        # Save the workbook
        wb.save(output_path)

        return {
            'status': 'success',
            'message': 'Invoice processed successfully',
            'output_path': output_path,
            'filename': output_filename
        }

    except Exception as e:
        print(f"Error processing files: {str(e)}")
        traceback.print_exc()
        return {
            'status': 'error',
            'message': str(e)
        }



# Route to get Excel sheet names - simplified version for initial deployment
@app.route('/api/get-sheets', methods=['POST'])
def get_excel_sheets():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
            
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.endswith(('.xls', '.xlsx')):
            return jsonify({'error': 'File must be an Excel file (.xls or .xlsx)'}), 400
            
        # For initial deployment, return dummy sheet names
        # This will be replaced with actual Excel parsing later
        dummy_sheet_names = ['Sheet1', 'Sheet2', 'Data']
        
        return jsonify({
            'status': 'success',
            'message': 'Full Excel parsing will be implemented in a future update',
            'sheet_names': dummy_sheet_names
        })
        
    except Exception as e:
        print(f"Error getting Excel sheets: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'error': f"An error occurred: {str(e)}"
        }), 500

# Route to download processed files - supports both path and query parameters
@app.route('/api/download-file', methods=['GET'])
@app.route('/api/download-file/<path:filename>', methods=['GET'])
def download_file(filename=None):
    try:
        # Check if filename is provided as query parameter (priority)
        query_filename = request.args.get('filename')
        if query_filename:
            filename = query_filename
        
        # If we still don't have a filename, return an error
        if not filename:
            return jsonify({
                'error': "No filename provided. Use /api/download-file/<filename> or /api/download-file?filename=<filename>"
            }), 400
        
        print(f"Attempting to download file: {filename}")
        
        # Ensure the filename is secure
        filename = secure_filename(filename)
        
        # Check if file exists
        file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
        if not os.path.exists(file_path):
            return jsonify({
                'error': f"File not found: {filename}"
            }), 404
            
        return send_from_directory(
            directory=app.config['PROCESSED_FOLDER'],
            path=filename,
            as_attachment=True
        )
    except Exception as e:
        print(f"Error downloading file: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'error': f"File not found or error downloading: {str(e)}"
        }), 404

if __name__ == '__main__':
    # Get port from environment variable or use default 10000
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_ENV') == 'development')
